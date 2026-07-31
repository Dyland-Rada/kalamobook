"""
Publicacion en Shopify: tabla de control, importador y generador de fichas.

Hoy la tienda se alimenta a mano: alguien genera un XLSX Matrixify y lo sube.
Este modulo automatiza las tres piezas:

  1. `shopify_productos`: las 23 columnas Matrixify de lo que YA esta en la
     tienda (734.957 ISBN el 31/07/2026). Es el control de duplicados, el
     diccionario de la taxonomia real y el banco de ejemplos para la IA.
  2. Generador de fichas: la ficha tecnica sale de nuestros datos y solo los
     bloques narrativos y el SEO los escribe DeepSeek.
  3. Salidas: XLSX Matrixify para la puesta al dia (52.773 pendientes) y API
     de Shopify para el goteo diario.

Por que separadas: Shopify limita a 1.000 variantes nuevas al dia cuando la
tienda pasa de 50.000, asi que la carga grande no puede ir por API.

Spec: docs/superpowers/specs/2026-07-31-shopify-publicacion-design.md
"""
import glob
import os
import time
from datetime import datetime

import db

TABLA = "shopify_productos"

# Cabeceras Matrixify -> columna en la tabla. El orden manda: es el mismo que
# se escribe al exportar el XLSX.
COLUMNAS = [
    ("Command", "command"),
    ("Handle", "handle"),
    ("Title", "title"),
    ("Vendor", "vendor"),
    ("Type", "tipo"),
    ("Tags", "tags"),
    ("Published", "published"),
    ("Status", "status"),
    ("Body HTML", "body_html"),
    ("SEO Title", "seo_title"),
    ("SEO Description", "seo_description"),
    ("Variant SKU", "variant_sku"),
    ("Variant Barcode", "variant_barcode"),
    ("Variant Price", "variant_price"),
    ("Variant Compare At Price", "variant_compare_at_price"),
    ("Variant Inventory Qty", "variant_inventory_qty"),
    ("Variant Inventory Tracker", "variant_inventory_tracker"),
    ("Variant Grams", "variant_grams"),
    ("Variant Requires Shipping", "variant_requires_shipping"),
    ("Image Src", "image_src"),
    ("Image Alt Text", "image_alt_text"),
    ("Metafield: custom.autor [single_line_text_field]", "metafield_autor"),
    ("Metafield: custom.anio_publicacion [number_integer]", "metafield_anio"),
]
CABECERAS = [c for c, _ in COLUMNAS]

_DDL = f"""
    CREATE TABLE IF NOT EXISTS {TABLA} (
        handle TEXT PRIMARY KEY,
        command TEXT,
        title TEXT,
        vendor TEXT,
        tipo TEXT,
        tags TEXT,
        published TEXT,
        status TEXT,
        body_html TEXT,
        seo_title TEXT,
        seo_description TEXT,
        variant_sku TEXT,
        variant_barcode TEXT,
        variant_price TEXT,
        variant_compare_at_price TEXT,
        variant_inventory_qty INTEGER,
        variant_inventory_tracker TEXT,
        variant_grams INTEGER,
        variant_requires_shipping TEXT,
        image_src TEXT,
        image_alt_text TEXT,
        metafield_autor TEXT,
        metafield_anio TEXT,
        estado TEXT NOT NULL DEFAULT 'publicado',
        fichero_origen TEXT,
        subido_en TIMESTAMP,
        cargado_en TIMESTAMP,
        generado_en TIMESTAMP
    )
"""
_INDICES = (
    f"CREATE INDEX IF NOT EXISTS {TABLA}_estado_idx ON {TABLA} (estado)",
)

_schema_ok = False


def ensure_schema():
    global _schema_ok
    if _schema_ok:
        return
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        cur.execute(_DDL)
        for sql in _INDICES:
            cur.execute(sql)
        conn.commit()
        _schema_ok = True
    except Exception as e:
        print(f"[ShopifyPub] ensure_schema FAIL: {e}")
        try: conn.rollback()
        except Exception: pass
    finally:
        conn.close()


# ─── Estado del job (para la tarjeta web) ────────────────────────────
_job: dict | None = None


def get_status() -> dict:
    job = dict(_job) if _job else {"status": "idle"}
    if "errors" in job:
        job["errors"] = job["errors"][-15:]
    return job


def stop() -> bool:
    if _job and _job.get("status") == "running":
        _job["status"] = "stopped"
        return True
    return False


# ─── Importar los XLSX ya subidos ────────────────────────────────────
def _int_o_none(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def importar_carpeta(carpeta: str, lote: int = 500) -> dict:
    """
    Carga los XLSX Matrixify de `carpeta` en la tabla. Los procesa en orden
    de fecha de fichero: si un ISBN aparece en varias tandas, gana la mas
    reciente, que es la que esta viva en la tienda.

    Lee en streaming (read_only) porque son ~870 MB y no caben en memoria.
    """
    global _job
    ensure_schema()
    import openpyxl
    from psycopg2.extras import execute_values

    ficheros = sorted(glob.glob(os.path.join(carpeta, "*.xlsx")),
                      key=lambda f: os.path.getmtime(f))
    _job = {
        "status": "running", "accion": "importar",
        "carpeta": carpeta, "ficheros": len(ficheros),
        "started_at": datetime.now().isoformat(), "stage": "empezando",
        "fichero_actual": None, "filas": 0, "insertadas": 0,
        "errors": [], "elapsed_s": 0,
    }
    job = _job
    t0 = time.monotonic()
    destino = [dest for _, dest in COLUMNAS]
    sql = f"""
        INSERT INTO {TABLA} ({", ".join(destino)},
                             estado, fichero_origen, subido_en, cargado_en)
        VALUES %s
        ON CONFLICT (handle) DO UPDATE SET
            {", ".join(f"{d}=EXCLUDED.{d}" for d in destino if d != "handle")},
            estado='publicado',
            fichero_origen=EXCLUDED.fichero_origen,
            subido_en=EXCLUDED.subido_en,
            cargado_en=NOW()
    """
    plantilla = "(" + ",".join(["%s"] * (len(destino) + 3)) + ", NOW())"

    conn = db.get_connection()
    cur = conn.cursor()
    try:
        for f in ficheros:
            if job["status"] != "running":
                break
            nombre = os.path.basename(f)
            fecha = datetime.fromtimestamp(os.path.getmtime(f))
            job["fichero_actual"] = nombre
            job["stage"] = f"leyendo {nombre}"
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            cabecera = list(next(it))
            idx = {h: i for i, h in enumerate(cabecera)}
            faltan = [c for c in CABECERAS if c not in idx]
            if faltan:
                job["errors"].append(f"{nombre}: faltan columnas {faltan[:3]}")
                wb.close()
                continue

            buffer = []
            i_handle = idx["Handle"]
            for r in it:
                # openpyxl en read_only devuelve la fila RECORTADA cuando las
                # ultimas celdas van vacias, asi que r puede ser mas corta que
                # la cabecera: hay que leer con red.
                if not r or len(r) <= i_handle or not r[i_handle]:
                    continue
                fila = []
                for cab, dest in COLUMNAS:
                    i = idx[cab]
                    v = r[i] if i < len(r) else None
                    if dest in ("variant_inventory_qty", "variant_grams"):
                        v = _int_o_none(v)
                    elif v is not None:
                        v = str(v).strip()
                    fila.append(v)
                buffer.append(tuple(fila) + ("publicado", nombre, fecha))
                job["filas"] += 1
                if len(buffer) >= lote:
                    execute_values(cur, sql, buffer, template=plantilla,
                                   page_size=lote)
                    conn.commit()
                    job["insertadas"] += len(buffer)
                    buffer = []
            if buffer:
                execute_values(cur, sql, buffer, template=plantilla,
                               page_size=len(buffer))
                conn.commit()
                job["insertadas"] += len(buffer)
            wb.close()
            print(f"[ShopifyPub] {nombre}: {job['insertadas']:,} acumuladas "
                  f"({time.monotonic() - t0:.0f}s)", flush=True)

        if job["status"] == "running":
            job["status"] = "completed"
        job["stage"] = "done"
    except Exception as e:
        job["status"] = "error"
        job["errors"].append(f"{type(e).__name__}: {e}"[:300])
        print(f"[ShopifyPub] importar FAIL: {e!r}")
        try: conn.rollback()
        except Exception: pass
    finally:
        conn.close()
        job["elapsed_s"] = round(time.monotonic() - t0, 1)
        _audit("importar",
               f"Importados {job['insertadas']:,} productos de Shopify desde "
               f"{job['ficheros']} ficheros ({job['elapsed_s']}s)", job,
               error=(job["status"] == "error"))
    return job


# ─── Consultas de estado ─────────────────────────────────────────────
def resumen() -> dict:
    """Cuantos hay publicados, generados y pendientes."""
    ensure_schema()
    conn = db.get_connection()
    cur = conn.cursor()
    out = {"publicados": 0, "generados": 0, "pendientes_bd": 0}
    try:
        db.execute_query(cur, f"SELECT estado, COUNT(*) FROM {TABLA} GROUP BY estado")
        por_estado = {r[0]: int(r[1]) for r in cur.fetchall()}
        out["publicados"] = por_estado.get("publicado", 0)
        out["generados"] = por_estado.get("generado", 0)
        out["por_estado"] = por_estado
    except Exception as e:
        out["error"] = str(e)[:150]
    finally:
        conn.close()
    return out


# Minimo para que una ficha no salga coja en la tienda. Medido sobre los
# 52.773 candidatos brutos del 31/07: 6.124 sin titulo (saldrian con el ISBN
# por nombre), 11.412 sin precio (no se pueden vender), 3.228 con una
# descripcion de una linea y 299 con un codigo interno que no es ISBN.
# Con estos filtros quedan 35.911 publicables de verdad.
DESC_MINIMA = int(os.environ.get("SHOPIFY_DESC_MINIMA", "120"))

# El precio va en cascada. `list_price` es el precio web ya calculado, pero
# esta vacio en 3.248 libros a los que el sync nunca se lo escribio; el PVP
# de verdad lo manda el proveedor en su CEGALD y Server A lo guarda en
# libros_proveedor.precio_con_iva. Sin esta cascada se quedaban fuera.
# El umbral de 2,90 es la regla API-15: por debajo no se publica.
_PRECIO_SQL = """
    COALESCE(NULLIF(m.list_price, 0), NULLIF(m.pvp_base, 0), prov.precio)
"""
_JOIN_PRECIO = """
    LEFT JOIN LATERAL (
        SELECT MAX(lp.precio_con_iva) AS precio
        FROM libros_proveedor lp
        WHERE lp.isbn = m.barcode AND lp.stock_disponible > 0
    ) prov ON true
"""

_SQL_CANDIDATOS = f"""
    FROM odoo_books_mirror m
    {_JOIN_PRECIO}
    WHERE m.barcode IS NOT NULL
      AND m.barcode ~ '^97[89]'
      AND NULLIF(m.name, '') IS NOT NULL
      AND m.name <> m.barcode
      AND COALESCE(NULLIF(m.cdl_image_url,''),
                   NULLIF(m.gbooks_thumbnail,'')) IS NOT NULL
      AND NULLIF(m.description,'') IS NOT NULL
      AND LENGTH(m.description) >= {DESC_MINIMA}
      AND {_PRECIO_SQL} >= 2.90
      AND EXISTS (SELECT 1 FROM libros_proveedor lp
                  WHERE lp.isbn = m.barcode AND lp.stock_disponible > 0)
      AND NOT EXISTS (SELECT 1 FROM {TABLA} s WHERE s.handle = m.barcode)
"""


def candidatos_sin_publicar(limite: int | None = None) -> list[str]:
    """
    ISBN que deberian estar en la tienda y no estan.

    Exige: ISBN de verdad (978/979), titulo propio, portada, descripcion con
    cuerpo, precio y stock. El peso NO filtra: si falta se estima en 350 g
    (regla del cliente), que es lo que distingue a los de etiqueta 2.
    """
    ensure_schema()
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        sql = f"SELECT m.barcode {_SQL_CANDIDATOS}"
        if limite:
            sql += f" LIMIT {int(limite)}"
        db.execute_query(cur, sql)
        return [r[0] for r in cur.fetchall()]
    finally:
        conn.close()


def candidatos_descartados() -> dict:
    """
    Por que se queda fuera cada candidato bruto. Para la tarjeta web: dice
    que hace falta conseguir para poder publicarlos.
    """
    ensure_schema()
    conn = db.get_connection()
    cur = conn.cursor()
    base = f"""
        FROM odoo_books_mirror m
        {_JOIN_PRECIO}
        WHERE m.barcode IS NOT NULL
          AND COALESCE(NULLIF(m.cdl_image_url,''),
                       NULLIF(m.gbooks_thumbnail,'')) IS NOT NULL
          AND NULLIF(m.description,'') IS NOT NULL
          AND EXISTS (SELECT 1 FROM libros_proveedor lp
                      WHERE lp.isbn = m.barcode AND lp.stock_disponible > 0)
          AND NOT EXISTS (SELECT 1 FROM {TABLA} s WHERE s.handle = m.barcode)
    """
    try:
        db.execute_query(cur, f"""
            SELECT COUNT(*),
                   COUNT(*) FILTER (WHERE NULLIF(m.name,'') IS NULL
                                       OR m.name = m.barcode),
                   COUNT(*) FILTER (WHERE COALESCE({_PRECIO_SQL}, 0) < 2.90),
                   COUNT(*) FILTER (WHERE LENGTH(m.description) < {DESC_MINIMA}),
                   COUNT(*) FILTER (WHERE m.barcode !~ '^97[89]'),
                   COUNT(*) FILTER (WHERE NULLIF(m.cdl_author,'') IS NULL)
            {base}
        """)
        r = cur.fetchone()
        return {"brutos": int(r[0]), "sin_titulo": int(r[1]),
                "sin_precio": int(r[2]), "descripcion_corta": int(r[3]),
                "sin_isbn": int(r[4]), "sin_autor": int(r[5])}
    except Exception as e:
        return {"error": str(e)[:150]}
    finally:
        conn.close()


def taxonomia() -> dict:
    """Las madres y categorias reales que usa la tienda, sacadas de lo subido."""
    ensure_schema()
    conn = db.get_connection()
    cur = conn.cursor()
    madres, cats = {}, {}
    try:
        db.execute_query(cur, f"""
            SELECT trim(t) AS etiqueta, COUNT(*)
            FROM {TABLA}, unnest(string_to_array(tags, ',')) AS t
            WHERE tags IS NOT NULL
            GROUP BY 1 ORDER BY 2 DESC
        """)
        for etiqueta, n in cur.fetchall():
            if etiqueta.startswith("madre:"):
                madres[etiqueta] = int(n)
            elif etiqueta.startswith("cat:"):
                cats[etiqueta] = int(n)
    except Exception as e:
        print(f"[ShopifyPub] taxonomia FAIL: {e}")
    finally:
        conn.close()
    return {"madres": madres, "cats": cats}


# ─── Auditoria contra la tienda ──────────────────────────────────────
def auditar(escribir: bool = True) -> dict:
    """
    Compara la tienda con nuestra tabla y la pone al dia.

    Los ficheros Matrixify no son toda la verdad: el 31/07 habia 11.971
    productos en Shopify que no venian de ahi (los mete la app de Odoo). Sin
    cruzar contra la tienda, 11.697 de nuestros "candidatos" ya estaban
    publicados y habriamos gastado ~34M de tokens en regenerarlos.

    Con escribir=True los que falten se anotan como publicados, para que la
    consulta de candidatos deje de proponerlos.
    """
    global _job
    import shopify_api as sa
    from psycopg2.extras import execute_values

    ensure_schema()
    _job = {
        "status": "running", "accion": "auditar",
        "started_at": datetime.now().isoformat(), "stage": "exportando de Shopify",
        "en_shopify": 0, "en_tabla": 0, "solo_en_shopify": 0,
        "solo_en_tabla": 0, "anotados": 0, "candidatos": 0,
        "errors": [], "elapsed_s": 0,
    }
    job = _job
    t0 = time.monotonic()
    try:
        productos = sa.exportar_handles()
        handles = {p["handle"] for p in productos if p.get("handle")}
        job["en_shopify"] = len(handles)

        conn = db.get_connection()
        cur = conn.cursor()
        try:
            db.execute_query(cur, f"SELECT handle FROM {TABLA}")
            tabla = {r[0] for r in cur.fetchall()}
            job["en_tabla"] = len(tabla)
            faltan = sorted(handles - tabla)
            job["solo_en_shopify"] = len(faltan)
            job["solo_en_tabla"] = len(tabla - handles)

            if escribir and faltan:
                job["stage"] = "anotando los que faltaban"
                execute_values(cur, f"""
                    INSERT INTO {TABLA} (handle, estado, fichero_origen, cargado_en)
                    VALUES %s
                    ON CONFLICT (handle) DO UPDATE
                    SET estado='publicado', cargado_en=NOW()
                """, [(h,) for h in faltan],
                    template="(%s, 'publicado', 'shopify-api', NOW())",
                    page_size=1000)
                conn.commit()
                job["anotados"] = len(faltan)
        finally:
            conn.close()

        job["candidatos"] = len(candidatos_sin_publicar())
        job["status"] = "completed"
        job["stage"] = "done"
    except Exception as e:
        job["status"] = "error"
        job["errors"].append(f"{type(e).__name__}: {e}"[:300])
        print(f"[ShopifyPub] auditar FAIL: {e!r}")
    finally:
        job["elapsed_s"] = round(time.monotonic() - t0, 1)
        _audit("auditar",
               f"Auditoria Shopify: {job['en_shopify']:,} en la tienda, "
               f"{job['en_tabla']:,} en la tabla, {job['solo_en_shopify']:,} solo "
               f"en la tienda ({job['anotados']:,} anotados), quedan "
               f"{job['candidatos']:,} por publicar ({job['elapsed_s']}s)", job,
               error=(job["status"] == "error"))
    return job


# ─── Generar fichas ──────────────────────────────────────────────────
CONCURRENCIA = int(os.environ.get("SHOPIFY_GEN_CONCURRENCIA", "10"))
DIR_SALIDA = os.environ.get("SHOPIFY_DIR", "reports")


def _guardar_fichas(filas: list[dict]):
    """Deja las fichas en la tabla como 'generado', listas para exportar."""
    from psycopg2.extras import execute_values
    destino = [dest for _, dest in COLUMNAS]
    sql = f"""
        INSERT INTO {TABLA} ({", ".join(destino)}, estado, generado_en, cargado_en)
        VALUES %s
        ON CONFLICT (handle) DO UPDATE SET
            {", ".join(f"{d}=EXCLUDED.{d}" for d in destino if d != "handle")},
            estado='generado', generado_en=NOW(), cargado_en=NOW()
    """
    plantilla = "(" + ",".join(["%s"] * len(destino)) + ", 'generado', NOW(), NOW())"
    valores = []
    for f in filas:
        fila = []
        for cab, dest in COLUMNAS:
            v = f.get(cab)
            if dest in ("variant_inventory_qty", "variant_grams"):
                v = _int_o_none(v)
            fila.append(v)
        valores.append(tuple(fila))
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        execute_values(cur, sql, valores, template=plantilla,
                       page_size=len(valores))
        conn.commit()
    finally:
        conn.close()


def generar_fichas(limite: int | None = None,
                   concurrencia: int | None = None,
                   lote_guardado: int = 50) -> dict:
    """
    Genera con IA las fichas de los candidatos y las guarda como 'generado'.

    NO publica nada: solo escribe en nuestra tabla. La subida a la tienda es
    otro paso (XLSX o API). Se guarda por lotes segun van saliendo, asi que
    una interrupcion no tira por la borda lo ya generado.
    """
    global _job
    import concurrent.futures as cf
    import shopify_ficha as sf

    ensure_schema()
    isbns = candidatos_sin_publicar(limite=limite)
    conc = concurrencia or CONCURRENCIA
    _job = {
        "status": "running", "accion": "generar",
        "started_at": datetime.now().isoformat(), "stage": "generando",
        "total": len(isbns), "generadas": 0, "fallidas": 0,
        "tokens_entrada": 0, "tokens_salida": 0, "llamadas": 0,
        "concurrencia": conc, "errors": [], "elapsed_s": 0,
    }
    job = _job
    t0 = time.monotonic()
    try:
        sf.taxonomia()          # cachear antes de repartir hilos
        pendientes = []
        with cf.ThreadPoolExecutor(max_workers=conc) as ex:
            futuros = {ex.submit(sf.generar, i): i for i in isbns}
            for fut in cf.as_completed(futuros):
                if job["status"] != "running":
                    break
                isbn = futuros[fut]
                try:
                    fila, uso = fut.result()
                    pendientes.append(fila)
                    job["generadas"] += 1
                    job["tokens_entrada"] += uso["prompt_tokens"]
                    job["tokens_salida"] += uso["completion_tokens"]
                    job["llamadas"] += uso["llamadas"]
                except Exception as e:
                    job["fallidas"] += 1
                    if len(job["errors"]) < 40:
                        job["errors"].append(f"{isbn}: {type(e).__name__}: {str(e)[:110]}")
                if len(pendientes) >= lote_guardado:
                    _guardar_fichas(pendientes)
                    pendientes = []
                    print(f"[ShopifyPub] {job['generadas']:,}/{job['total']:,} "
                          f"({time.monotonic() - t0:.0f}s)", flush=True)
        if pendientes:
            _guardar_fichas(pendientes)
        if job["status"] == "running":
            job["status"] = "completed"
        job["stage"] = "done"
    except Exception as e:
        job["status"] = "error"
        job["errors"].append(f"{type(e).__name__}: {e}"[:300])
        print(f"[ShopifyPub] generar FAIL: {e!r}")
    finally:
        job["elapsed_s"] = round(time.monotonic() - t0, 1)
        _audit("generar_fichas",
               f"Generadas {job['generadas']:,} fichas para Shopify "
               f"({job['fallidas']} fallidas, "
               f"{job['tokens_entrada'] + job['tokens_salida']:,} tokens, "
               f"{job['elapsed_s']}s)", job,
               error=(job["status"] == "error"))
    return job


# ─── Publicar en la tienda ───────────────────────────────────────────
def publicar(limite: int | None = None, dry_run: bool = True,
             canal_online: bool = True) -> dict:
    """
    Sube a Shopify las fichas que estan como 'generado'.

    Shopify corta en 1.000 variantes nuevas al dia por encima de 50.000, asi
    que hay tope. Lo que no entra hoy se queda como 'generado' y sale manana.
    Empezar SIEMPRE con dry_run=True: dice que subiria sin tocar la tienda.
    """
    global _job
    import shopify_api as sa

    ensure_schema()
    tope = min(limite or sa.TOPE_DIARIO, sa.TOPE_DIARIO)
    _job = {
        "status": "running", "accion": "publicar", "dry_run": dry_run,
        "started_at": datetime.now().isoformat(), "stage": "leyendo fichas",
        "disponibles": 0, "tope": tope, "publicados": 0, "fallidos": 0,
        "errors": [], "elapsed_s": 0,
    }
    job = _job
    t0 = time.monotonic()
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        db.execute_query(cur, f"SELECT COUNT(*) FROM {TABLA} WHERE estado='generado'")
        job["disponibles"] = int(cur.fetchone()[0])
        db.execute_query(cur, f"""
            SELECT {", ".join(d for _, d in COLUMNAS)}
            FROM {TABLA} WHERE estado='generado' ORDER BY handle LIMIT {tope}
        """)
        filas = [dict(zip(CABECERAS, r)) for r in cur.fetchall()]
        job["a_publicar"] = len(filas)

        if dry_run:
            job["muestra"] = [{"handle": f["Handle"], "title": f["Title"],
                               "tags": f["Tags"]} for f in filas[:20]]
            job["stage"] = "dry_run_done"
            job["status"] = "completed"
            return job

        canales_ids = []
        if canal_online:
            for c in sa.canales():
                if c["nombre"] in ("Online Store", "Shop"):
                    canales_ids.append(f"gid://shopify/Publication/{c['id']}")

        job["stage"] = "publicando"
        for fila in filas:
            if job["status"] != "running":
                break
            try:
                res = sa.crear_producto(fila, canales_ids or None)
                db.execute_query(cur, f"""
                    UPDATE {TABLA} SET estado='publicado', cargado_en=NOW()
                    WHERE handle = ?
                """, (fila["Handle"],))
                conn.commit()
                job["publicados"] += 1
                if job["publicados"] % 50 == 0:
                    print(f"[ShopifyPub] {job['publicados']}/{len(filas)} "
                          f"({time.monotonic() - t0:.0f}s)", flush=True)
            except Exception as e:
                job["fallidos"] += 1
                if len(job["errors"]) < 40:
                    job["errors"].append(f"{fila['Handle']}: {str(e)[:130]}")
        if job["status"] == "running":
            job["status"] = "completed"
        job["stage"] = "done"
    except Exception as e:
        job["status"] = "error"
        job["errors"].append(f"{type(e).__name__}: {e}"[:300])
        print(f"[ShopifyPub] publicar FAIL: {e!r}")
    finally:
        conn.close()
        job["elapsed_s"] = round(time.monotonic() - t0, 1)
        if not dry_run:
            _audit("publicar",
                   f"Publicados en Shopify {job['publicados']:,} productos "
                   f"({job['fallidos']} fallidos, {job['elapsed_s']}s)", job,
                   error=(job["status"] == "error" or job["fallidos"] > 0))
    return job


# ─── Exportar el XLSX Matrixify ──────────────────────────────────────
def exportar_xlsx(destino: str | None = None, estado: str = "generado",
                  limite: int | None = None) -> dict:
    """
    Escribe el XLSX que se sube con Matrixify, leyendo de la tabla.

    En modo write_only para no cargar el libro entero en memoria: son cientos
    de miles de filas con 4 KB de HTML cada una.
    """
    global _job
    import openpyxl

    ensure_schema()
    os.makedirs(DIR_SALIDA, exist_ok=True)
    if not destino:
        marca = datetime.now().strftime("%Y%m%d_%H%M")
        destino = os.path.join(DIR_SALIDA, f"Kalamo_Matrixify_{marca}.xlsx")
    _job = {
        "status": "running", "accion": "exportar",
        "started_at": datetime.now().isoformat(), "stage": "exportando",
        "destino": destino, "estado": estado, "filas": 0,
        "errors": [], "elapsed_s": 0,
    }
    job = _job
    t0 = time.monotonic()
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Products")
        ws.append(CABECERAS)
        sql = f"""SELECT {", ".join(d for _, d in COLUMNAS)}
                  FROM {TABLA} WHERE estado = ? ORDER BY handle"""
        if limite:
            sql += f" LIMIT {int(limite)}"
        db.execute_query(cur, sql, (estado,))
        while True:
            filas = cur.fetchmany(1000)
            if not filas:
                break
            for r in filas:
                ws.append(["" if v is None else v for v in r])
                job["filas"] += 1
        wb.save(destino)
        wb.close()
        job["tamano_mb"] = round(os.path.getsize(destino) / 1024 / 1024, 1)
        job["status"] = "completed"
        job["stage"] = "done"
    except Exception as e:
        job["status"] = "error"
        job["errors"].append(f"{type(e).__name__}: {e}"[:300])
        print(f"[ShopifyPub] exportar FAIL: {e!r}")
    finally:
        conn.close()
        job["elapsed_s"] = round(time.monotonic() - t0, 1)
        _audit("exportar_xlsx",
               f"Exportadas {job['filas']:,} fichas a {os.path.basename(destino)} "
               f"({job.get('tamano_mb', 0)} MB, {job['elapsed_s']}s)", job,
               error=(job["status"] == "error"))
    return job


def _audit(evento: str, resumen_txt: str, detalle: dict, error: bool = False):
    try:
        import audit_log
        audit_log.log_event("shopify", evento, resumen_txt,
                            detalle={k: v for k, v in detalle.items()
                                     if k != "errors"},
                            nivel="error" if error else "info")
    except Exception:
        pass


if __name__ == "__main__":
    import sys
    accion = sys.argv[1] if len(sys.argv) > 1 else "resumen"
    if accion == "importar":
        carpeta = sys.argv[2] if len(sys.argv) > 2 else "todos los libros webs"
        print(importar_carpeta(carpeta))
    elif accion == "resumen":
        print(resumen())
    elif accion == "taxonomia":
        t = taxonomia()
        print(f"{len(t['madres'])} madres, {len(t['cats'])} categorias")
