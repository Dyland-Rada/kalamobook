"""
Transforma nuevoslibros/Kalamo_COMPLETADOS.xlsx (8 columnas) al formato
Matrixify de 23 columnas para importar a Shopify.

Rellena lo que tenemos (Excel + imagen desde odoo_books_mirror) y deja
vacio lo que no (Body HTML y SEO Description -> vienen del ticket de
plantillas / copys IA). NO toca Odoo ni ninguna BD: solo lee el mirror
para las imagenes y escribe un xlsx nuevo.

Uso:
    DATABASE_URL=... python scripts/build_matrixify.py --limit 200 --out muestra.xlsx
    DATABASE_URL=... python scripts/build_matrixify.py            # completo
"""
import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

SRC = "nuevoslibros/Kalamo_COMPLETADOS.xlsx"
SHEET_IN = "Completados"

HEADERS = [
    "Command", "Handle", "Title", "Vendor", "Type", "Tags", "Published",
    "Status", "Body HTML", "SEO Title", "SEO Description", "Variant SKU",
    "Variant Barcode", "Variant Price", "Variant Compare At Price",
    "Variant Inventory Qty", "Variant Inventory Tracker", "Variant Grams",
    "Variant Requires Shipping", "Image Src", "Image Alt Text",
    "Metafield: custom.autor [single_line_text_field]",
    "Metafield: custom.anio_publicacion [number_integer]",
]


def clean_xml(s):
    """Quita caracteres de control invalidos en XML 1.0 (rompen el .xlsx).
    Validos: \\t \\n \\r, x20-D7FF, E000-FFFD, 10000-10FFFF."""
    if s is None:
        return ""
    out = []
    for ch in str(s):
        if ch in "\t\n\r" or ("\x20" <= ch <= "퟿") \
                or ("" <= ch <= "�") or ch >= "\U00010000":
            out.append(ch)
    return "".join(out).strip()


def price_str(p):
    if p is None or p == "":
        return ""
    try:
        return f"{float(str(p).replace(',', '.')):.2f}"
    except (ValueError, TypeError):
        return ""


def truncate(s, n):
    s = s or ""
    return s if len(s) <= n else s[:n]


def load_images():
    """barcode -> URL de imagen (cdl_image_url, fallback gbooks_thumbnail)."""
    conn = db.get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT barcode,
               COALESCE(NULLIF(cdl_image_url, ''), NULLIF(gbooks_thumbnail, ''))
        FROM odoo_books_mirror
        WHERE barcode IS NOT NULL
          AND (COALESCE(cdl_image_url, '') <> '' OR COALESCE(gbooks_thumbnail, '') <> '')
    """)
    out = {}
    for bc, url in cur.fetchall():
        if url:
            out[bc] = url
    conn.close()
    return out


def build(limit=None, out="nuevoslibros/Kalamo_MATRIXIFY.xlsx"):
    print("Cargando imagenes del mirror...")
    images = load_images()
    print(f"  {len(images):,} barcodes con imagen")

    wb_in = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws_in = wb_in[SHEET_IN]
    rows = ws_in.iter_rows(values_only=True)
    next(rows)  # cabecera origen

    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet("Products")
    ws_out.append(HEADERS)

    n = 0
    con_imagen = 0
    for r in rows:
        isbn = clean_xml(r[0])
        if not isbn:
            continue
        titulo = clean_xml(r[1])
        vendor = clean_xml(r[2]).upper()
        autor = clean_xml(r[3])
        anio = clean_xml(r[4])
        precio = price_str(r[5])
        grams = int(r[6]) if r[6] not in (None, "") else 0

        seo_title = truncate(f"{titulo} - {autor}" if autor else titulo, 60)
        alt = (f"{titulo}, de {autor} | portada del libro" if autor
               else f"{titulo} | portada del libro")
        alt = truncate(alt, 125)
        img = images.get(isbn, "")
        if img:
            con_imagen += 1

        ws_out.append([
            "MERGE",             # 1 Command
            isbn,                # 2 Handle
            titulo,              # 3 Title
            vendor,              # 4 Vendor
            "Libro",             # 5 Type
            "",                  # 6 Tags (taxonomia no disponible)
            "TRUE",              # 7 Published
            "active",            # 8 Status
            "",                  # 9 Body HTML (copys IA, ticket plantillas)
            seo_title,           # 10 SEO Title
            "",                  # 11 SEO Description (depende del copy)
            isbn,                # 12 Variant SKU
            isbn,                # 13 Variant Barcode
            precio,              # 14 Variant Price
            precio,              # 15 Variant Compare At Price
            0,                   # 16 Variant Inventory Qty
            "shopify",           # 17 Variant Inventory Tracker
            grams,               # 18 Variant Grams
            "TRUE",              # 19 Variant Requires Shipping
            img,                 # 20 Image Src
            alt if img else "",  # 21 Image Alt Text
            autor,               # 22 Metafield autor
            anio,                # 23 Metafield anio
        ])
        n += 1
        if n % 50000 == 0:
            print(f"  {n:,} filas escritas...")
        if limit and n >= limit:
            break

    wb_in.close()
    wb_out.save(out)
    print(f"LISTO: {n:,} filas -> {out}")
    print(f"  con imagen: {con_imagen:,} ({100*con_imagen/max(n,1):.0f}%)")
    print("  sin Body HTML ni SEO Description (pendiente ticket plantillas)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--out", default="nuevoslibros/Kalamo_MATRIXIFY.xlsx")
    a = ap.parse_args()
    build(limit=a.limit, out=a.out)
