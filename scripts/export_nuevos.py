"""
Exporta a Excel los libros NUEVOS que llegaron (los ~98k creados hoy en
Odoo, que no estaban antes). Incluye scrapeados y no scrapeados en el mismo
archivo, con etiqueta de completitud y proveedor.
"""
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db
from odoo_tags import _classify

OUT = "nuevoslibros/Libros_Nuevos_Llegados.xlsx"
RUBEN_FILE = "nuevoslibros/Kalamo_MATRIXIFY.xlsx"

HEADERS = ["ISBN", "Titulo", "Autor", "Editorial", "Precio", "Peso_g",
           "Imagen_URL", "Descripcion", "Proveedores", "Estado",
           "Etiqueta", "En_archivo_Ruben"]


def clean(s):
    if s is None:
        return ""
    out = []
    for ch in str(s):
        if ch in "\t\n\r" or ("\x20" <= ch <= "퟿") \
                or ("" <= ch <= "�") or ch >= "\U00010000":
            out.append(ch)
    return "".join(out).strip()


def load_ruben():
    wb = openpyxl.load_workbook(RUBEN_FILE, read_only=True, data_only=True)
    ws = wb["Products"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    s = {str(r[1]).strip() for r in rows if r[1]}
    wb.close()
    return s


def main():
    print("Cargando ISBNs del archivo de Ruben...")
    ruben = load_ruben()
    print(f"  {len(ruben):,}")

    conn = db.get_connection()
    cur = conn.cursor()
    print("Consultando libros nuevos (creados hoy)...")
    cur.execute("""
        SELECT m.barcode,
               b.title, b.author, b.editorial, b.image_url, b.description,
               b.weight, b.height, b.width,
               MAX(lp.precio_con_iva) AS precio,
               string_agg(DISTINCT p.nombre, ', ') AS proveedores
        FROM odoo_books_mirror m
        LEFT JOIN books b ON b.isbn = m.barcode
        LEFT JOIN libros_proveedor lp
               ON lp.isbn = m.barcode AND lp.stock_disponible > 0
        LEFT JOIN proveedores p ON p.id = lp.proveedor_id
        WHERE m.synced_at::date = CURRENT_DATE
        GROUP BY m.barcode, b.title, b.author, b.editorial, b.image_url,
                 b.description, b.weight, b.height, b.width
    """)
    rows = cur.fetchall()
    conn.close()
    print(f"  {len(rows):,} filas")

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Nuevos")
    ws.append(HEADERS)

    stats = {"scrap": 0, "no_scrap": 0}
    tag_counts = {}
    for (isbn, title, author, editorial, img, desc, weight, height, width,
         precio, prov) in rows:
        tag = _classify(img, desc, weight, height, width)
        tag_counts[tag] = tag_counts.get(tag, 0) + 1
        scraped = bool(title and str(title).strip())
        if scraped:
            stats["scrap"] += 1
        else:
            stats["no_scrap"] += 1
        ws.append([
            isbn,
            clean(title),
            clean(author),
            clean(editorial),
            float(precio) if precio is not None else "",
            int(weight) if weight and str(weight).strip().isdigit() else clean(weight),
            clean(img),
            clean(desc),
            clean(prov),
            "Scrapeado" if scraped else "No scrapeado",
            tag,
            "Si" if isbn in ruben else "No",
        ])

    wb.save(OUT)
    print(f"LISTO: {len(rows):,} filas -> {OUT}")
    print(f"  Scrapeados: {stats['scrap']:,} | No scrapeados: {stats['no_scrap']:,}")
    print(f"  Por etiqueta: {tag_counts}")


if __name__ == "__main__":
    main()
