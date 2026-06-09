"""
Import bulk de catalogos de distribuidores desde Excel a Postgres.

Cada archivo XLSX en `nuevoslibros/` viene con ~33k-100k libros y 36
columnas (mismo schema que la tabla `books` + `fuente`). Se upsertan por
ISBN a la tabla `distributor_books`. Idempotente.

Uso:
  1) Endpoint /api/v1/distributors/import (multipart file upload)
  2) Funcion `import_xlsx_from_path()` para scripts locales/server.
"""
import io
import os
from datetime import datetime
from typing import Any

import openpyxl

import db


# Orden EXACTO de columnas tal como aparecen en los XLSX.
# Cualquier columna extra/movida sera ignorada por nombre.
XLSX_FIELDS = [
    "id", "title", "author", "editorial", "isbn", "price", "original_price",
    "discount", "description", "translator", "illustrator", "language",
    "pages", "reading_time", "binding", "release_date", "edition_year",
    "edition_place", "collection", "height", "width", "weight", "origin",
    "url", "image_url", "category", "categoria_1", "categoria_2",
    "categoria_3", "timestamp", "price_eur", "sinli_situacion",
    "sinli_updated_at", "fuente", "categoria_4", "categoria_5",
]

# Solo las columnas que vamos a persistir (descartamos `id` y `timestamp`
# del XLSX porque son internos del archivo, no de nuestra BD).
DB_FIELDS = [
    "isbn", "title", "author", "editorial", "price", "original_price",
    "discount", "description", "translator", "illustrator", "language",
    "pages", "reading_time", "binding", "release_date", "edition_year",
    "edition_place", "collection", "height", "width", "weight", "origin",
    "url", "image_url", "category", "categoria_1", "categoria_2",
    "categoria_3", "categoria_4", "categoria_5", "price_eur",
    "sinli_situacion", "sinli_updated_at", "fuente",
]


# Estado para que la UI pueda monitorear progreso (un solo job a la vez).
import_job: dict | None = None


def get_import_status() -> dict:
    job = dict(import_job) if import_job else {"status": "idle"}
    job["total_in_db"] = count_distributor_rows()
    if "errors" in job:
        job["errors"] = job["errors"][-10:]
    return job


def count_distributor_rows() -> int:
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        db.execute_query(cur, "SELECT COUNT(*) FROM distributor_books")
        return cur.fetchone()[0]
    except Exception:
        return 0
    finally:
        conn.close()


def count_by_source() -> list[dict]:
    conn = db.get_connection()
    cur = conn.cursor()
    try:
        db.execute_query(cur, """
            SELECT COALESCE(fuente, '(sin fuente)') as src, COUNT(*) as n
            FROM distributor_books
            GROUP BY fuente
            ORDER BY n DESC
        """)
        return [{"fuente": r[0], "count": r[1]} for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        conn.close()


# ── Normalizers ─────────────────────────────────────────────────────────
def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _to_isbn(v: Any) -> str | None:
    """ISBN sale como int o str — siempre como str de digitos."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        s = str(int(v))
    else:
        s = str(v)
    digits = "".join(c for c in s if c.isdigit())
    return digits if len(digits) in (10, 13) else None


def _to_decimal(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


# ── Core import ────────────────────────────────────────────────────────
def _normalize_row(row: tuple, fuente_hint: str | None) -> dict | None:
    """Convierte una fila del XLSX al dict listo para upsert. None si invalida."""
    if not row or len(row) < 5:
        return None

    # Construir dict por nombre de columna
    d = {}
    for i, key in enumerate(XLSX_FIELDS):
        d[key] = row[i] if i < len(row) else None

    isbn = _to_isbn(d.get("isbn"))
    if not isbn:
        return None

    fuente = _to_str(d.get("fuente")) or fuente_hint

    return {
        "isbn": isbn,
        "title": _to_str(d.get("title")),
        "author": _to_str(d.get("author")),
        "editorial": _to_str(d.get("editorial")),
        "price": _to_decimal(d.get("price")),
        "original_price": _to_decimal(d.get("original_price")),
        "discount": _to_str(d.get("discount")),
        "description": _to_str(d.get("description")),
        "translator": _to_str(d.get("translator")),
        "illustrator": _to_str(d.get("illustrator")),
        "language": _to_str(d.get("language")),
        "pages": _to_str(d.get("pages")),
        "reading_time": _to_str(d.get("reading_time")),
        "binding": _to_str(d.get("binding")),
        "release_date": _to_str(d.get("release_date")),
        "edition_year": _to_str(d.get("edition_year")),
        "edition_place": _to_str(d.get("edition_place")),
        "collection": _to_str(d.get("collection")),
        "height": _to_str(d.get("height")),
        "width": _to_str(d.get("width")),
        "weight": _to_str(d.get("weight")),
        "origin": _to_str(d.get("origin")),
        "url": _to_str(d.get("url")),
        "image_url": _to_str(d.get("image_url")),
        "category": _to_str(d.get("category")),
        "categoria_1": _to_str(d.get("categoria_1")),
        "categoria_2": _to_str(d.get("categoria_2")),
        "categoria_3": _to_str(d.get("categoria_3")),
        "categoria_4": _to_str(d.get("categoria_4")),
        "categoria_5": _to_str(d.get("categoria_5")),
        "price_eur": _to_decimal(d.get("price_eur")),
        "sinli_situacion": _to_str(d.get("sinli_situacion")),
        "sinli_updated_at": _to_str(d.get("sinli_updated_at")),
        "fuente": fuente,
    }


def _upsert_batch(rows: list[dict]) -> int:
    if not rows:
        return 0
    conn = db.get_connection()
    cur = conn.cursor()
    inserted = 0

    cols = DB_FIELDS
    placeholders = ", ".join(["?"] * len(cols))
    col_names = ", ".join(cols)

    if db.IS_POSTGRES:
        # UPSERT por ISBN
        updates = ", ".join(f"{c} = EXCLUDED.{c}" for c in cols if c != "isbn")
        sql = (
            f"INSERT INTO distributor_books ({col_names}, imported_at) "
            f"VALUES ({placeholders}, CURRENT_TIMESTAMP) "
            f"ON CONFLICT (isbn) DO UPDATE SET {updates}, "
            f"imported_at = CURRENT_TIMESTAMP"
        )
    else:
        sql = (
            f"INSERT OR REPLACE INTO distributor_books "
            f"({col_names}, imported_at) "
            f"VALUES ({placeholders}, CURRENT_TIMESTAMP)"
        )

    for r in rows:
        try:
            db.execute_query(cur, sql, tuple(r.get(c) for c in cols))
            inserted += 1
        except Exception:
            # Una fila mala no debe abortar el lote
            continue

    conn.commit()
    conn.close()
    return inserted


def _import_workbook(wb, fuente_hint: str | None, batch_size: int = 500) -> dict:
    """
    Itera todas las hojas del workbook, normaliza filas y upserta en lotes.
    Devuelve metrics del import.
    """
    global import_job
    total_rows = 0
    inserted_total = 0
    skipped = 0
    errors = []
    batch: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Skip header row
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            total_rows += 1
            normalized = _normalize_row(row, fuente_hint)
            if not normalized:
                skipped += 1
                continue
            batch.append(normalized)

            if len(batch) >= batch_size:
                n = _upsert_batch(batch)
                inserted_total += n
                batch = []
                if import_job is not None:
                    import_job["processed"] = total_rows
                    import_job["inserted"] = inserted_total
                    import_job["skipped"] = skipped

    if batch:
        n = _upsert_batch(batch)
        inserted_total += n
        if import_job is not None:
            import_job["processed"] = total_rows
            import_job["inserted"] = inserted_total
            import_job["skipped"] = skipped

    return {
        "processed_rows": total_rows,
        "inserted_or_updated": inserted_total,
        "skipped_invalid_isbn": skipped,
        "errors": errors,
    }


# ── Public API ─────────────────────────────────────────────────────────
def import_xlsx_bytes(content: bytes, fuente_hint: str | None = None,
                      batch_size: int = 500) -> dict:
    """Lee XLSX desde bytes (upload) y upserta a Postgres."""
    global import_job
    import_job = {
        "status": "running",
        "started_at": datetime.now().isoformat(),
        "fuente_hint": fuente_hint,
        "processed": 0,
        "inserted": 0,
        "skipped": 0,
        "errors": [],
    }
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        result = _import_workbook(wb, fuente_hint, batch_size)
        wb.close()
        import_job["status"] = "completed"
        import_job["result"] = result
        return result
    except Exception as e:
        import_job["status"] = "error"
        err = f"{type(e).__name__}: {e!r}"
        import_job["errors"].append(err)
        raise


def import_xlsx_from_path(path: str, fuente_hint: str | None = None,
                          batch_size: int = 500) -> dict:
    """Lee XLSX desde un path en disco y upserta a Postgres."""
    global import_job
    fuente = fuente_hint or _guess_fuente_from_path(path)
    import_job = {
        "status": "running",
        "started_at": datetime.now().isoformat(),
        "path": path,
        "fuente_hint": fuente,
        "processed": 0,
        "inserted": 0,
        "skipped": 0,
        "errors": [],
    }
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result = _import_workbook(wb, fuente, batch_size)
        wb.close()
        import_job["status"] = "completed"
        import_job["result"] = result
        return result
    except Exception as e:
        import_job["status"] = "error"
        err = f"{type(e).__name__}: {e!r}"
        import_job["errors"].append(err)
        raise


def _guess_fuente_from_path(path: str) -> str | None:
    """Adivina el distribuidor desde el nombre del archivo."""
    name = os.path.basename(path).upper()
    for distrib in ("ANAYA", "PLANETA", "PODIPRINT", "LOGISTA"):
        if distrib in name:
            return distrib
    return None
